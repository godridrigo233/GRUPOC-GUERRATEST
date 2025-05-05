import tkinter as tk
from tkinter import messagebox
import sqlite3
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from datetime import datetime
from fpdf import FPDF
from PIL import Image, ImageTk

def crear_frame_reporte(root):
    frame_reporte = tk.Frame(root, bg="white")
    
    # Cargar y colocar el logo en la esquina superior derecha
    ruta_logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo.png')
    if os.path.exists(ruta_logo):
        try:
            # Cargar la imagen y redimensionarla si es necesario
            logo_image = Image.open(ruta_logo)
            logo_image = logo_image.resize((95, 95), Image.LANCZOS)  # Cambia el tamaño según sea necesario
            logo_photo = ImageTk.PhotoImage(logo_image)
            
            # Colocar el logo en la esquina superior derecha
            logo_label = tk.Label(frame_reporte, image=logo_photo, bg="white")
            logo_label.image = logo_photo  # Mantener referencia
            logo_label.place(relx=1.0, y=20, anchor="ne", x=-20)  # Ajusta 'y' y 'x' para desplazar hacia arriba y a la derecha
        except Exception as e:
            print(f"Error al cargar el logo: {e}")
    else:
        print("El archivo 'logo.png' no se encontró en la ruta especificada.")

    # Título
    titulo_label = tk.Label(
        frame_reporte, 
        text="DESCARGA EL REPORTE DE TU INVENTARIO", 
        font=("Arial", 35, "bold"), 
        bg="white"
    )
    titulo_label.pack(pady=40)
    
    # Espaciador para bajar los botones
    espaciador = tk.Label(frame_reporte, bg="white")
    espaciador.pack(pady=50)  # Ajusta el valor para obtener la distancia deseada
    
    # Cargar el ícono de Excel
    ruta_icono_excel = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel.png')
    if os.path.exists(ruta_icono_excel):
        try:
            icono_excel_img = Image.open(ruta_icono_excel)
            icono_excel_img = icono_excel_img.resize((80, 80), Image.LANCZOS)  # Cambia el tamaño según sea necesario
            icono_excel = ImageTk.PhotoImage(icono_excel_img)
        except Exception as e:
            print(f"Error al cargar el icono de Excel: {e}")
            icono_excel = None
    else:
        print("No se encontró el ícono de Excel.")
        icono_excel = None

    btn_descargar_excel = tk.Button(
        frame_reporte,
        text="Descargar Excel",
        image=icono_excel if icono_excel else None,
        compound="top",
        command=descargar_reporte_excel,
        font=("Arial", 20, "bold"),  # ⬆️ Texto más grande y negrita
        padx=20,     # ⬅️ espacio horizontal interno
        pady=20,     # ⬆️ espacio vertical interno
        bg="#1e7145",
        fg="white",
        bd=0,
        relief="flat",
        highlightthickness=0,
        activebackground="#155e38"
    )
    if icono_excel:
        btn_descargar_excel.image = icono_excel
    btn_descargar_excel.pack(pady=20)
    
        # Cargar el ícono de PDF
    ruta_icono_pdf = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pdf.png')
    if os.path.exists(ruta_icono_pdf):
        try:
            icono_pdf_img = Image.open(ruta_icono_pdf)
            icono_pdf_img = icono_pdf_img.resize((80, 80), Image.LANCZOS)
            icono_pdf = ImageTk.PhotoImage(icono_pdf_img)
        except Exception as e:
            print(f"Error al cargar el icono de PDF: {e}")
            icono_pdf = None
    else:
        print("No se encontró el ícono de PDF.")
        icono_pdf = None

    # Botón para descargar en PDF con diseño moderno
    btn_descargar_pdf = tk.Button(
        frame_reporte, 
        text="Descargar PDF", 
        image=icono_pdf if icono_pdf else None,
        compound="top",  # Ícono encima del texto
        command=descargar_reporte_pdf, 
        font=("Arial", 20, "bold"),
        padx=20,
        pady=20,
        bg="#C62828",
        fg="black",
        bd=0,
        relief="flat",
        highlightthickness=0,
        activebackground="#AD1C1C"
    )
    if icono_pdf:
        btn_descargar_pdf.image = icono_pdf  # mantener referencia

    btn_descargar_pdf.pack(pady=20)

    # Retorna el frame creado
    return frame_reporte

# Función para obtener la ruta de la base de datos
def get_database_path():
    # Detectar si estamos en el ejecutable creado por PyInstaller
    if getattr(sys, 'frozen', False):
        # Cuando se ejecuta como un ejecutable
        base_path = sys._MEIPASS
    else:
        # Cuando se ejecuta como script normal (por ejemplo, en VS Code)
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    # Retorna la ruta completa a ferreteria.db
    return os.path.join(base_path, "ferreteria.db")

def descargar_reporte_excel():
    # Conexión a la base de datos
    db_path = get_database_path()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Validar existencia de la tabla 'productos'
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='productos'")
    if not cursor.fetchone():
        conn.close()
        messagebox.showerror("Error", "La tabla 'productos' no existe en la base de datos.")
        return

    # Obtener datos de la tabla
    try:
        cursor.execute("SELECT * FROM productos")
        productos = cursor.fetchall()
        conn.close()
    except Exception as e:
        conn.close()
        messagebox.showerror("Error al consultar", str(e))
        return

    if not productos:
        messagebox.showinfo("Sin datos", "No hay registros en la base de datos para exportar.")
        return

    # Reconstruir columnas en orden deseado: ID, Nombre, Cantidad, Unidad, Precio
    columnas = ["ID", "Nombre", "Cantidad", "Unidad", "Precio"]
    data_ordenada = []
    for fila in productos:
        id_ = fila[0]
        nombre = fila[1]
        cantidad = fila[2]
        precio = round(fila[3], 2)  # Formatear a dos decimales
        unidad = fila[4]
        data_ordenada.append([id_, nombre, cantidad, unidad, precio])

    df = pd.DataFrame(data_ordenada, columns=columnas)

    # Crear carpeta si no existe
    ruta_carpeta_reporte = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Reporte")
    os.makedirs(ruta_carpeta_reporte, exist_ok=True)

    # Nombre de archivo con fecha y hora
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ruta_reporte = os.path.join(ruta_carpeta_reporte, f"Reporte_Inventario-{timestamp}.xlsx")

    # Exportar a Excel
    df.to_excel(ruta_reporte, index=False)

    # Ajustes de formato con openpyxl
    wb = load_workbook(ruta_reporte)
    ws = wb.active

    # Ajustar ancho automático
    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 2

    # Bordes finos para todas las celdas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    wb.save(ruta_reporte)
    messagebox.showinfo("Éxito", f"El reporte ha sido descargado exitosamente en {ruta_reporte}")


def descargar_reporte_pdf():
    db_path = get_database_path()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Validar existencia de la tabla 'productos'
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='productos'")
    if not cursor.fetchone():
        conn.close()
        messagebox.showerror("Error", "La tabla 'productos' no existe en la base de datos.")
        return

    # Obtener los datos
    try:
        cursor.execute("SELECT * FROM productos")
        productos = cursor.fetchall()
        conn.close()
    except Exception as e:
        conn.close()
        messagebox.showerror("Error al consultar", str(e))
        return

    if not productos:
        messagebox.showinfo("Sin datos", "No hay registros en la base de datos para exportar.")
        return

    columnas = ["ID", "Nombre", "Cantidad", "Unidad", "Precio"]

    ruta_carpeta_reporte = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Reporte")
    os.makedirs(ruta_carpeta_reporte, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fecha_actual = datetime.now().strftime("%d/%m/%y")
    hora_actual = datetime.now().strftime("%H:%M:%S")
    ruta_reporte_pdf = os.path.join(ruta_carpeta_reporte, f"Reporte_Inventario-{timestamp}.pdf")

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Logo
    ruta_logo = 'logo.png'
    if os.path.exists(ruta_logo):
        try:
            logo_img = Image.open(ruta_logo).convert("RGB")
            logo_img.save("temp_logo.jpg")
            pdf.image("temp_logo.jpg", x=186, y=5, w=16)
            os.remove("temp_logo.jpg")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el logo: {e}")

    # Título
    pdf.set_font("Arial", "B", 22)
    pdf.cell(0, 10, "REPORTE DE INVENTARIO - RUPHA", ln=True, align="C")
    pdf.ln(5)

    # Fecha y hora
    pdf.set_font("Arial", "", 10)
    pdf.set_text_color(128, 128, 128)
    pdf.cell(0, 10, f"Archivo creado: Fecha: {fecha_actual}    Hora: {hora_actual}", ln=True, align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(10)

    # Tabla
    pdf.set_font("Arial", "B", 12)
    col_width = 180 / 5
    row_height_default = 10

    def render_encabezados():
        pdf.set_font("Arial", "B", 12)
        for col in columnas:
            pdf.cell(col_width, row_height_default, col, border=1, align='C')
        pdf.ln(row_height_default)
        pdf.set_font("Arial", "", 10)

    render_encabezados()

    for fila in productos:
        # Orden: ID, Nombre, Cantidad, Unidad, Precio
        id_ = str(fila[0])
        nombre = str(fila[1])
        cantidad = str(fila[2])
        precio = f"{fila[3]:.2f}"
        unidad = str(fila[4])

        x_start = pdf.get_x()
        y_start = pdf.get_y()

        # Altura del nombre
        pdf.set_xy(x_start + col_width, y_start)
        nombre_lines = pdf.multi_cell(col_width, 5, nombre, border=0, align='C', split_only=True)
        row_height = max(len(nombre_lines) * 5, row_height_default)

        # Verificar salto de página
        if y_start + row_height > pdf.h - 15:
            pdf.add_page()
            render_encabezados()
            x_start = pdf.get_x()
            y_start = pdf.get_y()

        # Bordes manuales
        for i in range(5):
            pdf.rect(x_start + i * col_width, y_start, col_width, row_height)

        # Celdas
        pdf.set_xy(x_start, y_start)
        pdf.cell(col_width, row_height, id_, border=0, align='C')

        pdf.set_xy(x_start + col_width, y_start)
        pdf.multi_cell(col_width, 5, nombre, border=0, align='C')

        pdf.set_xy(x_start + col_width * 2, y_start)
        pdf.cell(col_width, row_height, cantidad, border=0, align='C')

        pdf.set_xy(x_start + col_width * 3, y_start)
        pdf.cell(col_width, row_height, unidad, border=0, align='C')

        pdf.set_xy(x_start + col_width * 4, y_start)
        pdf.cell(col_width, row_height, precio, border=0, align='C')

        pdf.set_y(y_start + row_height)

    try:
        pdf.output(ruta_reporte_pdf)
        messagebox.showinfo("Éxito", f"El reporte PDF ha sido descargado exitosamente en {ruta_reporte_pdf}")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo crear el archivo PDF: {e}")
